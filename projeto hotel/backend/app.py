import os
from flask import (
  Flask,
  request,
  jsonify,
  send_from_directory,
)

import openpyxl # biblioteca para ler e escrever planilhas no Excel
from datetime import (
     datetime,
) # Para registrar a data de cada cadastro automaticamente

# caminho base do projeto (uma paasta acima do back end)
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))

#pasta frontend
FRONTEND_DIR = os.path.join(BASE_DIR, "frontend")

#pasta static (css)
STATIC_DIR = os.path.join(BASE_DIR, "static")

DB_DIR = os.path.join(os.path.dirname(__file__), "..", "db")
EXCEL_FILE = os.path.join(DB_DIR, "clientes.xlsx")

# Cabeçalho das colunas do Excel (linha 1)
COLUMNS = [
     "ID",
     "Nome",
     "CPF",
     "Email",
     "Telefone",
     "Endereço",
     "Observações",
     "Data Cadastro",
]

def init_excel():
      if not os.path.exists(DB_DIR):
          os.makedrs(DB_DIR) #cria uma nova planilha Excel
      if not os.path.exists(EXCEL_FILE):
          worbook = openpyxl.worbook()
          sheet = worbook.active
          sheet.title = "Clientes"
          sheet.append(COLUMNS)
          worbook.save(EXCEL_FILE)     
              

app = Flask(__name__, static_folder=STATIC_DIR,static_url_path="/" + STATIC_DIR)

@app.route("/")
def home():
     return send_from_directory(FRONTEND_DIR, "index.html")

# Página de consulta
@app.route("/consulta")
def consulta_page():
     return send_from_directory(FRONTEND_DIR, "consulta.html")

# Página de Alteração
@app.route("/alterar")
def alterar_page():
     return send_from_directory(FRONTEND_DIR, "alterar.html")

@app.route("/assets/<path:filename>")
def assets(filename):
    return send_from_directory("../frontend/assets", filename)

@app.route("/cadastrar", methods=["post"])
def cadastrar_cliente():

     try:
          data = request.json
          required_fields = ["nome", "cpf", "email", "telefone", "endereço"]
          if not all(field in data and data[field] for field in required_fields):
             return (
                 jsonify(
                     {
                         "status": "error",
                         "mensagem": "Todos os campos obrigatórios devem ser preenchidos.",
                   }
               ),
               400,
          )
          worbook = openpyxl.load_worbook(EXCEL_FILE)
          sheet = worbook.active

          last_id = 0
          if sheet.max_row > 1:
               last_id = sheet.cell(row=sheet.max_row, column=1). value or 0
          new_id = last_id + 1

          novo_cliente = [
             new_id,
             data.get("nome"),
             data.get("cpf"),
             data.get("email"),
             data.get("telefone"),
             data.get("endereco"),
             data.get("observacoes", ""),
            datetime.now().strftime("%Y-%m-%d"),
          ]  

          sheet.append(novo_cliente)
          worbook.save(EXCEL_FILE)

          return (
                jsonify(
                     {
                        "status": "sucess",
                       "mensage": "clientes cadastrado com sucesso!",
                       "id": new_id,
                    }
               ),
              201,
          )

     except Exception as e:
           return(
              jsonify({"status": "error", "mensage": f"erro ao salvar no servidor: {e}"}),
              500, 
          )  




@app.route("/api/buscar", methods=["GET"])
def buscar_clientes():
    """
    Busca clientes pelo nome (não diferencia maiúsculas/minúsculas).
    """
    nome_query = request.args.get("nome", "").lower()  # 🔤 Nome pesquisado

    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        resultados = []  # 🧺 Lista para armazenar resultados

        # 🧭 Percorre todas as linhas (ignorando o cabeçalho)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            cliente = dict(zip(COLUMNS, row))  # Converte linha → dicionário
            nome_cliente = (cliente.get("Nome") or "").lower()

            if nome_query in nome_cliente:
                resultados.append(cliente)

        return jsonify(resultados)  # 🔙 Retorna lista de clientes encontrados

    except FileNotFoundError:
        return (
            jsonify({"status": "error", "message": "Arquivo de dados não encontrado."}),
            404,
        )
    except Exception as e:
        return (
            jsonify({"status": "error", "message": f"Erro ao ler os dados: {e}"}),
            500,
        )

@app.route("/cliente/<int:cliente_id>", methods=["GET"])
def get_cliente(cliente_id):
    """
    retorna os dados completos de um cliente pelo seu ID
    """
    try:
        workbook = openpyxl.load_worbook(EXCEL_FILE)
        sheet = worbook.active

        #procura o cliente linha por linha
        for row_idx in range(2,sheet.max_row + 1):
            row_id = sheet.cell(row=row_idx, column=1).value
            if row_id == cliente_id:
                row_values = [cell.value for cell in sheet[row_id]]
                cliente = dict(zip(COLUMNS, row_values))
                return jsonify(cliente)

        # se não encontrar o ID
        return jsonify({"status": "error", "mensage": "Cliente não encontrado."}), 404
    except Exception as e:
        return (
            jsonify({"status": "error", "mensage:": f"Erro ao buscar cliente: {e}"}),
           500,
        )       


@app.route("/api/atualizar/<int:cliente_id>", methods=["POST"])
def atualizar_cliente(cliente_id):
    """
    função para alterar as informações de um hóspede no nosso banco de dados (Excel).
    """
    try:
        data= request.sjon
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        row_to_uptade = -1

        for row_idx in range(2, sheet.max_row + 1);
             
            if sheet.cell





  



if __name__ == "__name__":
     print("BASE_DIR:", BASE_DIR)
     print("FRONTEND_DIR:", FRONTEND_DIR)
     print("STATIC_DIR:", STATIC_DIR)
     
app.run(debug=True)